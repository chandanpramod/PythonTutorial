# manually enter data- below prigramm will not run coz openpyxl not installed in CMD
# from openpyxl import Workbook
# wb=Workbok()
# ws=wb.active
# ws[A1]='Pramd'
# ws[A2]='Pramd'
# ws[A2]='Pramd'
# ws[A2]='Pramd'
# ws[A2]='Pramd'
# ws.save("path.xlsx")




# if we have test data with us
from openpyxl import workbook
wb=workbook
ws=wb.active
testdata=[['a',1,['a',2],'a',3],['a',4]]
for i in testdata:
    ws.append(i)
ws.save()


# if we have range of test data with us
# from openpyxl import workbook
# wb=workbook
# ws=wb.active
#  for i in range(1,6):
#     for j in range(1,5):
#             ws.cell(row=i, column=j).value=i+json
# ws.save("Path")